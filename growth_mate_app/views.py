import string
from django.contrib.sites.shortcuts import get_current_site
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import EmailMessage, send_mail
from django.template.loader import render_to_string
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import User
import random
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import Course, UserProfile, Lesson, DashboardStats, StudentProgress, Enrollment, CourseCategory, CourseTag, ContentBlock, CourseProgress, Activity
from .tokens import generate_token  
from growth_mate_project import settings 
from django.http import HttpResponseRedirect, JsonResponse
from django.core.mail import send_mail
import random
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Avg, Sum, Q
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.hashers import make_password
from django.utils import timezone
from social_django.utils import load_strategy, load_backend
from social_core.backends.oauth import BaseOAuth2PKCE
from social_core.exceptions import MissingBackend
from collections import defaultdict
import logging
from openpyxl import Workbook
from django.http import HttpResponse
from datetime import timedelta
from django.db import transaction
import random
from .gemini_service import get_bot_response
from .models import ChatMessage
from django.contrib.auth.tokens import default_token_generator

otp_storage = {}
logger = logging.getLogger(__name__)

def is_admin(user):
    return user.is_authenticated and (user.is_superuser or (hasattr(user, 'userprofile') and user.userprofile.role == 'admin'))

def signup(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password1")  # Changed from password to password1
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        role = request.POST.get("role")

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already registered.")
            return redirect("signup")

        # Generate OTP
        otp = ''.join(random.choices(string.digits, k=6))
        
        # Store user data and OTP in session
        request.session['temp_user_data'] = {
            'email': email,
            'password': password,
            'first_name': first_name,
            'last_name': last_name,
            'role': role
        }
        request.session['otp'] = otp
        request.session.set_expiry(300)  # 5 minutes expiry

        # Send OTP email
        send_mail(
            'Verify your email',
            f'Your OTP is: {otp}',
            settings.EMAIL_HOST_USER,
            [email],
            fail_silently=False,
        )

        messages.success(request, "Registration successful! Please check your email for OTP verification.")
        return redirect("verify_otp")

    return render(request, "signup.html")

def home(request):
    context = {}
    if request.user.is_authenticated:
        try:
            user_profile = UserProfile.objects.get(user=request.user)
            context['user_profile'] = user_profile
        except UserProfile.DoesNotExist:
            pass
    return render(request, "index.html", context)

def login_view(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        # Debug print
        print(f"Attempting login with email: {email}")

        # Try to get the user by email first
        try:
            user = User.objects.get(email=email)
            print(f"Found user: {user.username}, is_superuser: {user.is_superuser}")
        except User.DoesNotExist:
            print(f"No user found with email: {email}")
            messages.error(request, "Invalid login credentials, please try again.")
            return redirect("login")

        # Authenticate with username (which is email) and password
        user = authenticate(request, username=user.username, password=password)
        print(f"Authentication result: {'Success' if user else 'Failed'}")

        if user is None:
            messages.error(request, "Invalid login credentials, please try again.")
            return redirect("login")

        login(request, user)
        messages.success(request, "Login successful!")
        
        # Check user role and redirect accordingly
        try:
            user_profile = UserProfile.objects.get(user=user)
            print(f"User profile found with role: {user_profile.role}")
            if user.is_superuser:
                return redirect('admin_dashboard')
            elif user_profile.role == 'manager':
                return redirect('manager_dashboard')
            elif user_profile.role == 'employee':
                return redirect('employee_dashboard')
            else:
                return redirect('home')
        except UserProfile.DoesNotExist:
            print("No user profile found")
            # Create a user profile if it doesn't exist
            UserProfile.objects.create(user=user, role='admin' if user.is_superuser else 'employee')
            return redirect('admin_dashboard' if user.is_superuser else 'home')

    return render(request, "login.html")

def logout_view(request):
    user_profile = None
    is_admin = False
    if request.user.is_authenticated:
        try:
            user_profile = UserProfile.objects.get(user=request.user)
            is_admin = user_profile.role == 'admin' or request.user.is_superuser
        except UserProfile.DoesNotExist:
            is_admin = request.user.is_superuser
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    if is_admin or (user_profile and user_profile.role == 'manager'):
        return redirect("home")
    return redirect("login")

def verify_otp(request):
    if request.method == "POST":
        email = request.session.get('temp_user_data', {}).get('email')
        entered_otp = request.POST.get("otp")
        stored_otp = request.session.get('otp')

        if not email or not stored_otp:
            messages.error(request, "Session expired. Please register again.")
            return redirect("signup")

        if stored_otp == entered_otp:
            user_data = request.session.get('temp_user_data')

             # Create user
            user = User.objects.create_user(
                username=email,
                email=email,
                password=user_data['password'],
                first_name=user_data['first_name'],
                last_name=user_data['last_name']
            )

            # Update or create user profile
            UserProfile.objects.update_or_create(
                user=user,
                defaults={
                    'role': user_data['role']
                }
            )

            # Clear session data
            request.session.pop('temp_user_data', None)

            messages.success(request, "Your account has been activated! You can now log in.")
            return redirect("login")
        else:
            messages.error(request, "Invalid OTP. Please try again.")
            return redirect("verify_otp")

    return render(request, "verify_otp.html")


def resend_otp(request):
    if request.method == "POST":
        new_otp = random.randint(100000, 999999)  

        request.session['otp'] = new_otp  

        send_mail(
            "Your New OTP",
            f"Your OTP code is: {new_otp}",
            "no-reply@growthmate.com",
            [request.user.email], 
            fail_silently=False,
        )

        return JsonResponse({"message": "OTP has been resent!", "success": True})
    return JsonResponse({"message": "Invalid request", "success": False}, status=400)

def my_courses_view(request):
    trending_courses = Course.objects.all().order_by('-due_date')[:6]
    return render(request, "my_courses.html", {"trending_courses": trending_courses})

@login_required
def manager_dashboard(request):
    # Check if the user is a manager
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.role != 'manager':
            messages.error(request, "Access denied. Manager privileges required.")
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect('home')

    # Get today's date and last month's date
    today = timezone.now().date()
    last_month = today - timezone.timedelta(days=30)

    # Get or create today's statistics
    today_stats, created = DashboardStats.objects.get_or_create(date=today)
    last_month_stats = DashboardStats.objects.filter(date=last_month).first()

    # Calculate total users (employees only)
    total_users = User.objects.filter(userprofile__role='employee').count()
    active_users = User.objects.filter(userprofile__role='employee', is_active=True).count()
    inactive_users = total_users - active_users

    # Get manager's courses
    manager_courses = Course.objects.filter(instructor=request.user)
    total_courses = manager_courses.count()
    active_courses = manager_courses.filter(is_active=True).count()
    
    # Calculate course completion rate based on all enrolled users' progress
    total_progress = 0
    total_enrollments = 0
    
    for course in manager_courses:
        enrollments = Enrollment.objects.filter(course=course)
        for enrollment in enrollments:
            progress = CourseProgress.objects.filter(
                user=enrollment.user,
                course=course
            ).first()
            
            if progress:
                total_lessons = course.lessons.count()
                if total_lessons > 0:
                    completed_lessons = progress.completed_lessons.count()
                    enrollment_progress = (completed_lessons / total_lessons) * 100
                    total_progress += enrollment_progress
                    total_enrollments += 1
    
    course_completion = round(total_progress / total_enrollments, 2) if total_enrollments > 0 else 0

    # Calculate growth rates
    if last_month_stats:
        user_growth = round(((total_users - last_month_stats.total_users) / last_month_stats.total_users * 100) if last_month_stats.total_users > 0 else 0, 2)
        course_growth = round(((total_courses - last_month_stats.total_courses) / last_month_stats.total_courses * 100) if last_month_stats.total_courses > 0 else 0, 2)
        completion_growth = round((course_completion - last_month_stats.course_completion_rate), 2)
    else:
        user_growth = 0
        course_growth = 0
        completion_growth = 0

    # Update today's statistics
    today_stats.total_users = total_users
    today_stats.total_courses = total_courses
    today_stats.active_courses = active_courses
    today_stats.course_completion_rate = course_completion
    today_stats.user_growth_rate = user_growth
    today_stats.course_growth_rate = course_growth
    today_stats.completion_growth_rate = completion_growth
    today_stats.save()

    # Get top students based on progress
    top_students = []
    student_progress_records = StudentProgress.objects.filter(
        course__in=manager_courses
    ).select_related('user').order_by('-progress_percentage')[:5]
    
    for record in student_progress_records:
        top_students.append({
            'name': record.user.get_full_name() or record.user.username,
            'progress': record.progress_percentage
        })

    # Get course completion data with progress for all enrolled users
    course_completion_data = []
    for course in manager_courses:
        enrollments = Enrollment.objects.filter(course=course)
        total_course_progress = 0
        total_course_enrollments = 0
        
        for enrollment in enrollments:
            progress = CourseProgress.objects.filter(
                user=enrollment.user,
                course=course
            ).first()
            
            if progress:
                total_lessons = course.lessons.count()
                if total_lessons > 0:
                    completed_lessons = progress.completed_lessons.count()
                    enrollment_progress = (completed_lessons / total_lessons) * 100
                    total_course_progress += enrollment_progress
                    total_course_enrollments += 1
        
        course_progress = round(total_course_progress / total_course_enrollments, 2) if total_course_enrollments > 0 else 0
        course_completion_data.append({
            'title': course.title,
            'progress': course_progress
        })
    
    context = {
        'user_profile': user_profile,
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'total_courses': total_courses,
        'active_courses': active_courses,
        'course_completion': course_completion,
        'user_growth': user_growth,
        'course_growth': course_growth,
        'completion_growth': completion_growth,
        'top_students': top_students,
        'course_completion_data': course_completion_data
    }
    
    return render(request, 'manager/dashboard.html', context)

@login_required
def my_courses(request):
    user_profile = get_object_or_404(UserProfile, user=request.user)
    
    if user_profile.role == 'manager':
        # Get courses created by the manager
        created_courses = Course.objects.filter(instructor=request.user).order_by('-created_at')
        
        # Get courses the manager is enrolled in
        enrolled_courses = Course.objects.filter(
            enrollments__user=request.user
        ).exclude(
            instructor=request.user
        ).order_by('-created_at')
        
        # Combine both lists
        courses = list(created_courses) + list(enrolled_courses)
        
        # Add a flag to distinguish between created and enrolled courses
        for course in created_courses:
            course.is_created_by_me = True
        for course in enrolled_courses:
            course.is_created_by_me = False
            
        template = 'manager/my_courses.html'
    else:  # employee
        # Get courses through the enrollments related name
        enrolled_courses = Course.objects.filter(
            enrollments__user=request.user
        ).order_by('-created_at')
        courses = enrolled_courses
        template = 'employee/my_courses.html'
    
    # Add additional course data
    for course in courses:
        course.enrolled_students_count = course.enrollments.count()
        course.active_enrollments_count = course.enrollments.filter(completed=False).count()
        
        # Set is_enrolled attribute
        course.is_enrolled = True  # Since we're already filtering for enrolled courses
        
        # Calculate actual progress based on completed lessons
        if user_profile.role == 'employee' or (user_profile.role == 'manager' and not getattr(course, 'is_created_by_me', False)):
            # Get course progress
            course_progress = CourseProgress.objects.filter(
                user=request.user,
                course=course
            ).first()
            
            # Get total lessons count
            total_lessons = course.lessons.count()
            
            if course_progress and total_lessons > 0:
                # Calculate completion rate based on completed lessons
                completed_count = course_progress.completed_lessons.count()
                completion_rate = int((completed_count / total_lessons) * 100)
                
                # Update enrollment progress
                enrollment = Enrollment.objects.get(user=request.user, course=course)
                enrollment.progress = completion_rate
                if completion_rate == 100:
                    enrollment.completed = True
                else:
                    enrollment.completed = False
                enrollment.save()
                
                # Update StudentProgress for overall tracking
                student_progress, _ = StudentProgress.objects.get_or_create(
                    user=request.user,
                    course=course
                )
                student_progress.progress_percentage = completion_rate
                student_progress.save()
                
                # Add completion rate to course object for template
                course.user_completion_rate = completion_rate
            else:
                course.user_completion_rate = 0
        else:
            # For courses created by the manager, calculate average completion rate
            total_progress = 0
            total_enrollments = 0
            
            for enrollment in course.enrollments.all():
                progress = CourseProgress.objects.filter(
                    user=enrollment.user,
                    course=course
                ).first()
                
                if progress:
                    total_lessons = course.lessons.count()
                    if total_lessons > 0:
                        completed_lessons = progress.completed_lessons.count()
                        enrollment_progress = (completed_lessons / total_lessons) * 100
                        total_progress += enrollment_progress
                        total_enrollments += 1
            
            course.user_completion_rate = round(total_progress / total_enrollments, 2) if total_enrollments > 0 else 0
    
    return render(request, template, {
        'courses': courses,
        'user_profile': user_profile
    })

@login_required
@require_POST
def add_course(request):
    user_profile = get_object_or_404(UserProfile, user=request.user)
    if user_profile.role != 'manager':
        messages.error(request, 'Access denied. Managers only.')
        return redirect('home')
    
    try:
        course = Course.objects.create(
            title=request.POST['title'],
            duration=request.POST['duration'],
            due_date=request.POST['due_date'],
            about_this_course=request.POST['about_this_course'],
            uploaded_by=request.user,
            is_active='is_active' in request.POST
        )
        
        if 'image' in request.FILES:
            course.image = request.FILES['image']
            course.save()
            
        messages.success(request, 'Course created successfully!')
        return redirect('manager_courses')
    except Exception as e:
        messages.error(request, f'Error creating course: {str(e)}')
        return redirect('manager_courses')

@login_required
@require_POST
def add_lesson(request):
    user_profile = get_object_or_404(UserProfile, user=request.user)
    if user_profile.role != 'manager':
        messages.error(request, 'Access denied. Managers only.')
        return redirect('home')
    
    course = get_object_or_404(Course, id=request.POST['course_id'], uploaded_by=request.user)
    
    try:
        # Get the highest order number for this course's lessons
        last_order = Lesson.objects.filter(course=course).order_by('-order').first()
        new_order = (last_order.order + 1) if last_order else 1
        
        lesson = Lesson.objects.create(
            course=course,
            title=request.POST['title'],
            content=request.POST['content'],
            duration=request.POST['duration'],
            order=new_order
        )
        
        messages.success(request, 'Lesson added successfully!')
        return redirect('manager_courses')
    except Exception as e:
        messages.error(request, f'Error adding lesson: {str(e)}')
        return redirect('manager_courses')

@login_required
def edit_course(request, course_id):
    user_profile = get_object_or_404(UserProfile, user=request.user)
    if user_profile.role != 'manager':
        messages.error(request, 'Access denied. Managers only.')
        return redirect('home')
    
    course = get_object_or_404(Course, id=course_id, instructor=request.user)
    
    if request.method == 'POST':
        try:
            course.title = request.POST['title']
            course.duration = request.POST['duration']
            course.due_date = request.POST['due_date']
            course.about_this_course = request.POST['about_this_course']
            course.is_active = 'is_active' in request.POST
            
            if 'image' in request.FILES:
                course.image = request.FILES['image']
            
            course.save()
            messages.success(request, 'Course updated successfully!')
            return redirect('manager_courses')
        except Exception as e:
            messages.error(request, f'Error updating course: {str(e)}')
            return redirect('manager_courses')
    
    return render(request, 'manager/course_form.html', {
        'course': course,
        'user_profile': user_profile
    })

@login_required
def view_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    lessons = course.lesson_set.all().order_by('order')
    
    return render(request, 'view_course.html', {
        'course': course,
        'lessons': lessons
    })

@login_required
def profile_settings(request):
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        user_profile = UserProfile.objects.create(user=request.user)

    if request.method == 'POST':
        # Handle file uploads
        if 'profile_picture' in request.FILES:
            print("Profile picture uploaded:", request.FILES['profile_picture'])
            user_profile.profile_picture = request.FILES['profile_picture']
            user_profile.save()
            messages.success(request, 'Profile picture updated successfully.')
            return redirect('profile_settings')
            
        if 'cover_image' in request.FILES:
            print("Cover image uploaded:", request.FILES['cover_image'])
            user_profile.cover_image = request.FILES['cover_image']
            user_profile.save()
            messages.success(request, 'Cover image updated successfully.')
            return redirect('profile_settings')

        # Update user fields
        request.user.first_name = request.POST.get('first_name', request.user.first_name)
        request.user.last_name = request.POST.get('last_name', request.user.last_name)
        request.user.save()

        # Update profile fields
        if user_profile.role in ['manager', 'admin']:
            user_profile.professional_headline = request.POST.get('professional_headline', user_profile.professional_headline)
        user_profile.bio = request.POST.get('bio', user_profile.bio)
        user_profile.phone = request.POST.get('phone', user_profile.phone)
        user_profile.gender = request.POST.get('gender', user_profile.gender)
        user_profile.country = request.POST.get('country', user_profile.country)
        user_profile.city = request.POST.get('city', user_profile.city)
        user_profile.save()

        # Handle password change
        current_password = request.POST.get('current_password')
        if current_password:
            if request.user.check_password(current_password):
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')
                if new_password and confirm_password:
                    if new_password == confirm_password:
                        request.user.set_password(new_password)
                        request.user.save()
                        messages.success(request, 'Profile and password updated successfully.')
                        # Re-authenticate the user to prevent logout
                        update_session_auth_hash(request, request.user)
                    else:
                        messages.error(request, 'New passwords do not match.')
                elif new_password or confirm_password:
                    messages.error(request, 'Please provide both new password and confirmation.')
            else:
                messages.error(request, 'Current password is incorrect.')
        else:
            messages.success(request, 'Profile updated successfully.')

        return redirect('profile_settings')

    context = {
        'user_profile': user_profile,
    }
    
    # Use different templates based on user role
    if user_profile.role == 'admin':
        return render(request, 'admin/profile_settings.html', context)
    elif user_profile.role == 'manager':
        return render(request, 'manager/profile_settings.html', context)
    else:
        return render(request, 'employee/profile_settings.html', context)

@login_required
def users_view(request):
    # Check if the user is a manager
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.role != 'manager':
            messages.error(request, "Access denied. Manager privileges required.")
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect('home')

    # Get search query and filters
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'all')
    sort_by = request.GET.get('sort', 'name')  # Default sort by name
    sort_order = request.GET.get('order', 'asc')  # Default ascending order
    page = request.GET.get('page', 1)

    # Base queryset for employees
    employees = User.objects.filter(userprofile__role='employee')

    # Apply search filter
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    # Apply status filter
    if status_filter == 'active':
        employees = employees.filter(is_active=True)
    elif status_filter == 'inactive':
        employees = employees.filter(is_active=False)

    # Apply sorting
    sort_field = {
        'name': 'first_name',
        'date': 'date_joined'
    }.get(sort_by, 'first_name')

    if sort_order == 'desc':
        sort_field = f'-{sort_field}'
    employees = employees.order_by(sort_field)

    # Get manager's courses
    manager_courses = Course.objects.filter(instructor=request.user)
    
    # Calculate statistics
    total_users = User.objects.filter(userprofile__role='employee').count()
    active_users = User.objects.filter(userprofile__role='employee', is_active=True).count()
    inactive_users = total_users - active_users

    # Calculate course completion rate based on all enrolled users' progress
    total_progress = 0
    total_enrollments = 0
    
    for course in manager_courses:
        enrollments = Enrollment.objects.filter(course=course)
        for enrollment in enrollments:
            progress = CourseProgress.objects.filter(
                user=enrollment.user,
                course=course
            ).first()
            
            if progress:
                total_lessons = course.lessons.count()
                if total_lessons > 0:
                    completed_lessons = progress.completed_lessons.count()
                    enrollment_progress = (completed_lessons / total_lessons) * 100
                    total_progress += enrollment_progress
                    total_enrollments += 1
    
    course_completion = round(total_progress / total_enrollments, 2) if total_enrollments > 0 else 0

    # Get user activity data (last completed lessons)
    user_activities = []
    for employee in employees:
        # Get the most recent course progress for this employee
        last_completion = CourseProgress.objects.filter(
            user=employee,
            course__in=manager_courses,
            completed_lessons__isnull=False
        ).order_by('-last_accessed').first()
        
        if last_completion:
            # Get the most recently completed lesson for this course
            last_lesson = last_completion.completed_lessons.first()
            if last_lesson:
                user_activities.append({
                    'user': employee,
                    'lesson': last_lesson,
                    'course': last_lesson.course,
                    'completed_at': last_completion.last_accessed
                })
    
    # Sort activities by completion date
    user_activities.sort(key=lambda x: x['completed_at'], reverse=True)

    # Get course distribution data
    course_distribution = []
    for course in manager_courses:
        enrollment_count = course.enrollments.count()
        course_distribution.append({
            'title': course.title,
            'count': enrollment_count
        })

    # Paginate the results
    paginator = Paginator(employees, 10)  # Show 10 users per page
    try:
        employees = paginator.page(page)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)

    # Get additional user data
    for employee in employees:
        # Get enrolled courses count
        employee.enrolled_courses_count = Enrollment.objects.filter(
            user=employee,
            course__in=manager_courses
        ).count()
        
        # Get completed courses count
        employee.completed_courses_count = Enrollment.objects.filter(
            user=employee,
            course__in=manager_courses,
            completed=True
        ).count()
        
        # Get last login
        employee.last_login = employee.last_login.strftime('%Y-%m-%d %H:%M') if employee.last_login else 'Never'

    context = {
        'user_profile': user_profile,
        'employees': employees,
        'stats': {
            'total_users': total_users,
            'active_users': active_users,
            'inactive_users': inactive_users,
            'course_completion': course_completion
        },
        'search_query': search_query,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'page_obj': employees,
        'user_activities': user_activities[:10],  # Show last 10 activities
        'course_distribution': course_distribution
    }
    
    return render(request, 'manager/users.html', context)

@login_required
def courses_view(request):
    # Check if user is a manager
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.role != 'manager':
            messages.error(request, 'Access Denied: Manager privileges required.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found.')
        return redirect('home')

    # Static data for courses page
    context = {
        'user_profile': user_profile,
        'stats': {
            'total_courses': 56,
            'total_courses_trend': '+12%',
            'archive_courses': 14,
            'archive_courses_trend': '0%',
            'draft_pending': 33,
            'draft_pending_trend': '-5%',
            'enrollments': 22,
            'enrollments_trend': '+8%',
        },
        'active_courses_count': 42,
        'course_hours': 320,
    }

    return render(request, 'manager/courses.html', context)

@login_required
def select_role(request):
    if request.method == 'POST':
        role = request.POST.get('role')
        user_id = request.session.get('user_id')
        
        if user_id and role:
            user = User.objects.get(id=user_id)
            
            # Create user profile with social data if available
            UserProfile.objects.create(
                user=user,
                role=role,
                email=request.session.get('social_email', user.email),
                first_name=request.session.get('social_first_name', user.first_name),
                last_name=request.session.get('social_last_name', user.last_name)
            )
            
            # Clean up session
            for key in ['user_id', 'social_email', 'social_first_name', 'social_last_name']:
                request.session.pop(key, None)
            
            messages.success(request, 'Profile created successfully!')
            
            # Redirect based on role
            if role == 'manager':
                return redirect('manager_dashboard')
            elif role == 'employee':
                return redirect('employee_dashboard')
            else:
                return redirect('home')
        else:
            messages.error(request, 'Invalid role selection.')
            
    return render(request, 'select_role.html')

@login_required
def employee_dashboard(request):
    # Check if the user is an employee
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.role != 'employee':
            messages.error(request, "Access denied. Employee privileges required.")
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect('home')

    # Get enrolled courses
    enrolled_courses = Course.objects.filter(
        enrollments__user=request.user
    ).order_by('-created_at')
    
    # Calculate overall progress based on lesson completion across all courses
    total_progress = 0
    courses_with_progress = 0
    
    for course in enrolled_courses:
        progress = CourseProgress.objects.filter(
            user=request.user,
            course=course
        ).first()
        
        if progress:
            total_lessons = course.lessons.count()
            if total_lessons > 0:
                completed_lessons = progress.completed_lessons.count()
                course_progress = (completed_lessons / total_lessons) * 100
                total_progress += course_progress
                courses_with_progress += 1
    
    overall_progress = round(total_progress / courses_with_progress, 2) if courses_with_progress > 0 else 0
    
    # Get active courses (not completed)
    active_courses = enrolled_courses.filter(enrollments__user=request.user, enrollments__completed=False).count()
    
    # Calculate learning time (sum of course durations)
    total_learning_time = enrolled_courses.aggregate(total_time=Sum('duration'))['total_time'] or 0
    learning_time_hours = round(total_learning_time / 60, 1)  # Convert minutes to hours
    
    # Get recent course completions for achievements
    recent_completions = Enrollment.objects.filter(
        user=request.user,
        completed=True,
        completed_at__gte=timezone.now() - timezone.timedelta(days=30)
    ).count()
    
    # Get course progress data for the summary section
    course_progress_data = []
    for course in enrolled_courses[:5]:  # Get top 5 courses
        progress = CourseProgress.objects.filter(
            user=request.user,
            course=course
        ).first()
        
        if progress:
            total_lessons = course.lessons.count()
            if total_lessons > 0:
                completed_lessons = progress.completed_lessons.count()
                progress_percentage = round((completed_lessons / total_lessons) * 100, 2)
            else:
                progress_percentage = 0
        else:
            progress_percentage = 0
            
        course_progress_data.append({
            'title': course.title,
            'progress': progress_percentage
        })
    
    # Get next session (next incomplete lesson)
    next_lesson = None
    for course in enrolled_courses:
        progress = CourseProgress.objects.filter(
            user=request.user,
            course=course
        ).first()
        
        if progress:
            completed_lessons = progress.completed_lessons.all()
            next_lesson = course.lessons.exclude(id__in=completed_lessons).order_by('order').first()
            if next_lesson:
                break
    
    # Get pending tasks (incomplete lessons)
    pending_tasks = 0
    for course in enrolled_courses:
        progress = CourseProgress.objects.filter(
            user=request.user,
            course=course
        ).first()
        
        if progress:
            total_lessons = course.lessons.count()
            completed_lessons = progress.completed_lessons.count()
            pending_tasks += (total_lessons - completed_lessons)
    
    context = {
        'user_profile': user_profile,
        'overall_progress': overall_progress,
        'total_courses': enrolled_courses.count(),
        'completed_courses': enrolled_courses.filter(enrollments__user=request.user, enrollments__completed=True).count(),
        'active_courses': active_courses,
        'learning_time_hours': learning_time_hours,
        'recent_completions': recent_completions,
        'course_progress_data': course_progress_data,
        'next_lesson': next_lesson,
        'pending_tasks': pending_tasks,
    }
    return render(request, 'employee/dashboard.html', context)

@login_required
def available_courses(request):
    user_profile = get_object_or_404(UserProfile, user=request.user)
    
    # Get all active courses that the user is not enrolled in
    enrolled_course_ids = Course.objects.filter(
        enrollments__user=request.user
    ).values_list('id', flat=True)
    
    available_courses = Course.objects.filter(
        is_active=True
    ).exclude(
        id__in=enrolled_course_ids
    ).order_by('-created_at')
    
    return render(request, 'employee/available_courses.html', {
        'courses': available_courses,
        'user_profile': user_profile
    })

@login_required
def manager_courses(request):
    if not request.user.userprofile.role == 'manager':
        messages.error(request, 'Access denied. Only managers can access this page.')
        return redirect('home')
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '')
    category_filter = request.GET.get('category', 'all')
    status_filter = request.GET.get('status', 'all')
    
    # Get courses for the logged-in instructor
    courses = Course.objects.filter(instructor=request.user)
    
    # Apply filters
    if search_query:
        courses = courses.filter(title__icontains=search_query)
    if category_filter != 'all':
        courses = courses.filter(category_id=category_filter)
    if status_filter != 'all':
        courses = courses.filter(is_active=(status_filter == 'active'))
    
    # Calculate statistics
    stats = {
        'total_courses': courses.count(),
        'active_courses': courses.filter(is_active=True).count(),
        'total_enrollments': Enrollment.objects.filter(course__in=courses).count(),
        'completion_rate': 0
    }
    
    # Calculate overall completion rate for stats
    total_progress = 0
    total_enrollments = 0
    
    for course in courses:
        enrollments = Enrollment.objects.filter(course=course)
        for enrollment in enrollments:
            progress = CourseProgress.objects.filter(
                user=enrollment.user,
                course=course
            ).first()
            
            if progress:
                total_lessons = course.lessons.count()
                if total_lessons > 0:
                    completed_lessons = progress.completed_lessons.count()
                    enrollment_progress = (completed_lessons / total_lessons) * 100
                    total_progress += enrollment_progress
                    total_enrollments += 1
    
    stats['completion_rate'] = round(total_progress / total_enrollments, 2) if total_enrollments > 0 else 0
    
    # Add additional course data
    for course in courses:
        course.enrolled_students_count = course.enrollments.count()
        course.active_enrollments_count = course.enrollments.filter(completed=False).count()
        
        # Calculate course-specific lesson completion rate
        course_total_progress = 0
        course_total_enrollments = 0
        
        enrollments = course.enrollments.all()
        for enrollment in enrollments:
            progress = CourseProgress.objects.filter(
                user=enrollment.user,
                course=course
            ).first()
            
            if progress:
                total_lessons = course.lessons.count()
                if total_lessons > 0:
                    completed_lessons = progress.completed_lessons.count()
                    enrollment_progress = (completed_lessons / total_lessons) * 100
                    course_total_progress += enrollment_progress
                    course_total_enrollments += 1
        
        # Use a temporary attribute instead of completion_rate
        course.avg_lesson_completion_rate = round(course_total_progress / course_total_enrollments, 2) if course_total_enrollments > 0 else 0
    
    # Pagination
    paginator = Paginator(courses, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'user_profile': request.user.userprofile,
        'courses': page_obj,
        'stats': stats,
        'search_query': search_query,
        'category_filter': category_filter,
        'status_filter': status_filter,
    }
    return render(request, 'manager/courses.html', context)

def _handle_course_form(request, course_id=None, is_admin=False):
    """
    Base function to handle course form submission for both admin and manager roles.
    """
    course = None
    if course_id:
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            messages.error(request, 'Course not found.')
            return redirect('admin_courses' if is_admin else 'manager_courses')

    if request.method == 'POST':
        # Handle course data
        title = request.POST.get('title')
        description = request.POST.get('description')
        category_id = request.POST.get('category')
        level = request.POST.get('level')
        duration = request.POST.get('duration')
        prerequisites = request.POST.get('prerequisites')
        objectives = request.POST.get('objectives')
        target_audience = request.POST.get('target_audience')
        max_students = request.POST.get('max_students')
        is_active = request.POST.get('is_active') == 'on'
        is_featured = request.POST.get('is_featured') == 'on'
        certificate_available = request.POST.get('certificate_available') == 'on'

        # Handle thumbnail
        thumbnail = request.FILES.get('thumbnail')
        if thumbnail and course and course.thumbnail:
            # Delete old thumbnail if it exists
            course.thumbnail.delete()

        # Create or update course
        if not course:
            course = Course.objects.create(
                title=title,
                description=description,
                category_id=category_id,
                level=level,
                duration=duration,
                prerequisites=prerequisites,
                objectives=objectives,
                target_audience=target_audience,
                max_students=max_students,
                is_active=is_active,
                is_featured=is_featured,
                certificate_available=certificate_available,
                thumbnail=thumbnail,
                instructor=request.user
            )
        else:
            course.title = title
            course.description = description
            course.category_id = category_id
            course.level = level
            course.duration = duration
            course.prerequisites = prerequisites
            course.objectives = objectives
            course.target_audience = target_audience
            course.max_students = max_students
            course.is_active = is_active
            course.is_featured = is_featured
            course.certificate_available = certificate_available
            if thumbnail:
                course.thumbnail = thumbnail
            course.save()

        # Handle lessons
        lesson_titles = request.POST.getlist('lesson_title[]')
        lesson_durations = request.POST.getlist('lesson_duration[]')
        
        # Delete existing lessons if updating
        if course_id:
            course.lessons.all().delete()

        # Create new lessons
        for i in range(len(lesson_titles)):
            if lesson_titles[i] and lesson_durations[i]:
                lesson = Lesson.objects.create(
                    course=course,
                    title=lesson_titles[i],
                    duration=lesson_durations[i],
                    order=i
                )

                # Handle content blocks for this lesson
                content_blocks = request.POST.getlist(f'lesson_{i}_content_blocks[]')
                content_types = request.POST.getlist(f'lesson_{i}_content_types[]')
                content_files = request.FILES.getlist(f'lesson_{i}_content_files[]')

                for j in range(len(content_blocks)):
                    if content_blocks[j]:
                        content_type = content_types[j]
                        content = content_blocks[j]
                        file = content_files[j] if j < len(content_files) else None

                        ContentBlock.objects.create(
                            lesson=lesson,
                            content_type=content_type,
                            content=content,
                            file=file,
                            order=j
                        )

        messages.success(request, f'Course {"updated" if course_id else "created"} successfully.')
        return redirect('admin_courses' if is_admin else 'manager_courses')

    # Get categories for the form
    categories = CourseCategory.objects.all()

    context = {
        'course': course,
        'categories': categories,
    }
    return context

@user_passes_test(is_admin)
def admin_course_form(request, course_id=None):
    if request.user.userprofile.role != 'admin':
        return redirect('home')
    
    context = _handle_course_form(request, course_id, is_admin=True)
    if isinstance(context, HttpResponseRedirect):
        return context
    
    return render(request, 'admin/course_form.html', context)

@login_required
def course_form(request, course_id=None):
    # Check if user is a manager
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.role != 'manager':
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found.')
        return redirect('home')

    context = _handle_course_form(request, course_id, is_admin=False)
    if isinstance(context, HttpResponseRedirect):
        return context
    
    return render(request, 'manager/course_form.html', context)

@login_required
def delete_course(request, course_id):
    # Check if the user is a manager
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.role != 'manager':
            messages.error(request, "Access denied. Manager privileges required.")
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect('home')

    course = get_object_or_404(Course, id=course_id, instructor=request.user)
    
    if request.method == 'POST':
        course.delete()
        messages.success(request, 'Course deleted successfully.')
        return JsonResponse({'success': True, 'message': 'Course deleted successfully.'})
    
    # If it's an AJAX request but not POST, return course details for the modal
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'course': {
                'id': course.id,
                'title': course.title,
                'description': course.description
            }
        })
    
    # For non-AJAX GET requests, redirect to courses page
    return redirect('manager_courses')

@login_required
def enroll_course(request, course_id):
    user_profile = get_object_or_404(UserProfile, user=request.user)
    if user_profile.role not in ['employee', 'manager']:
        messages.error(request, 'Only employees and managers can enroll in courses.')
        return redirect('home')
    
    course = get_object_or_404(Course, id=course_id, is_active=True)
    
    # Check if user is already enrolled
    if Enrollment.objects.filter(user=request.user, course=course).exists():
        messages.warning(request, 'You are already enrolled in this course.')
        return redirect('my_courses')
    
    # Check if course has reached maximum students
    if course.max_students and course.enrollments.count() >= course.max_students:
        messages.error(request, 'This course has reached its maximum number of students.')
        return redirect('available_courses' if user_profile.role == 'employee' else 'manager_available_courses')
    
    try:
        # Create enrollment
        Enrollment.objects.create(
            user=request.user,
            course=course,
            enrolled_at=timezone.now()
        )
        messages.success(request, f'Successfully enrolled in {course.title}!')
        return redirect('my_courses')
    except Exception as e:
        messages.error(request, f'Error enrolling in course: {str(e)}')
        return redirect('available_courses' if user_profile.role == 'employee' else 'manager_available_courses')

@login_required
def manage_lessons(request, course_id):
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.role not in ['manager', 'admin']:
            messages.error(request, 'You do not have permission to manage lessons.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found.')
        return redirect('home')

    # Get course based on user role
    if user_profile.role == 'admin':
        course = get_object_or_404(Course, id=course_id)
    else:
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
    
    lessons = course.lessons.all().prefetch_related('content_blocks').order_by('order')

    if request.method == 'POST':
        try:
            # Log incoming request data
            logger.debug("=== POST Request Data ===")
            logger.debug("POST data: %s", dict(request.POST))
            logger.debug("FILES data: %s", dict(request.FILES))

            # Parse lesson data using defaultdict
            lesson_data = defaultdict(dict)
            for key in request.POST:
                if key.startswith('lessons-'):
                    parts = key.split('-', 2)
                    if len(parts) == 3 and parts[1].isdigit():
                        index = parts[1]
                        field = parts[2]
                        lesson_data[index][field] = request.POST[key]

            # Process each lesson
            for index, data in lesson_data.items():
                lesson_id = data.get('id')
                title = data.get('title')
                duration = data.get('duration')
                is_deleted = data.get('deleted') == 'true'

                logger.debug("Processing lesson index %s: id=%s, title=%s, duration=%s, deleted=%s",
                            index, lesson_id, title, duration, is_deleted)

                if lesson_id and is_deleted:
                    lesson = Lesson.objects.get(id=lesson_id, course=course)
                    lesson.delete()
                    logger.debug("Deleted lesson %s", lesson_id)
                    continue

                if lesson_id:
                    # Update existing lesson
                    lesson = Lesson.objects.get(id=lesson_id, course=course)
                    lesson.title = title
                    lesson.duration = duration
                    lesson.order = int(index)
                    lesson.save()
                    logger.debug("Updated lesson %s: title=%s, duration=%s", lesson_id, title, duration)
                else:
                    # Create new lesson
                    lesson = Lesson.objects.create(
                        course=course,
                        title=title,
                        duration=duration,
                        order=int(index)
                    )
                    logger.debug("Created new lesson %s: title=%s, duration=%s", lesson.id, title, duration)

                # Parse content blocks for this lesson
                content_block_data = defaultdict(dict)
                for key in request.POST:
                    if key.startswith(f'lessons-{index}-content_blocks-'):
                        parts = key.split('-', 4)
                        if len(parts) == 5 and parts[3].isdigit():
                            block_index = parts[3]
                            field = parts[4]
                            content_block_data[block_index][field] = request.POST[key]

                # Handle files for content blocks
                for block_index in content_block_data.keys():
                    file_key = f'lessons-{index}-content_blocks-{block_index}-file'
                    if file_key in request.FILES:
                        content_block_data[block_index]['file'] = request.FILES[file_key]

                # Process content blocks
                for block_index, block_data in content_block_data.items():
                    block_id = block_data.get('id')
                    text = block_data.get('text')
                    content_type = block_data.get('type')
                    file = block_data.get('file')
                    is_deleted = block_data.get('deleted') == 'true'

                    logger.debug("Processing content block index %s: id=%s, type=%s, text=%s, file=%s, deleted=%s",
                                block_index, block_id, content_type, text, file, is_deleted)

                    if block_id and is_deleted:
                        content_block = ContentBlock.objects.get(id=block_id, lesson=lesson)
                        content_block.delete()
                        logger.debug("Deleted content block %s", block_id)
                        continue

                    if block_id:
                        # Update existing content block
                        content_block = ContentBlock.objects.get(id=block_id, lesson=lesson)
                        content_block.content = text
                        content_block.content_type = content_type
                        if file:
                            content_block.file = file
                        content_block.order = int(block_index)
                        content_block.save()
                        logger.debug("Updated content block %s: type=%s, content=%s, file=%s",
                                    block_id, content_type, text, file)
                    else:
                        # Create new content block
                        if text or file:
                            content_block = ContentBlock.objects.create(
                                lesson=lesson,
                                content_type=content_type,
                                content=text,
                                file=file,
                                order=int(block_index)
                            )
                            logger.debug("Created content block %s: type=%s, content=%s, file=%s",
                                        content_block.id, content_type, text, file)

            messages.success(request, 'Lessons updated successfully!')
            # Redirect based on user role
            if user_profile.role == 'admin':
                return redirect('admin_courses')
            else:
                return redirect('manager_courses')
        except Exception as e:
            logger.error("Error in manage_lessons POST: %s", str(e))
            messages.error(request, f'Error updating lessons: {str(e)}')
            # Redirect based on user role
            if user_profile.role == 'admin':
                return redirect('admin_courses')
            else:
                return redirect('manager_courses')

    # Log lessons data before rendering
    logger.debug("Rendering lessons for course %s", course_id)
    for lesson in lessons:
        logger.debug("Lesson %s: %s", lesson.id, lesson.title)
        for block in lesson.content_blocks.all():
            logger.debug("  Content block %s: type=%s, content=%s, file=%s",
                        block.id, block.content_type, block.content, block.file)

    context = {
        'course': course,
        'lessons': lessons,
        'user_profile': user_profile,
    }
    
    # Use admin template for admin users, manager template for managers
    template = 'admin/manage_lessons.html' if user_profile.role == 'admin' else 'manager/manage_lessons.html'
    return render(request, template, context)

@login_required
def delete_lesson(request, course_id, lesson_id):
    """
    Delete a specific lesson from a course.
    """
    # Check if user is admin or manager
    user_profile = get_object_or_404(UserProfile, user=request.user)
    if user_profile.role not in ['admin', 'manager']:
        messages.error(request, 'Access denied. Admin or manager privileges required.')
        return redirect('home')
    
    # Get the course
    if user_profile.role == 'admin':
        course = get_object_or_404(Course, id=course_id)
    else:
        course = get_object_or_404(Course, id=course_id, instructor=request.user)
    
    # Get the lesson
    lesson = get_object_or_404(Lesson, id=lesson_id, course=course)
    
    # Delete the lesson
    lesson.delete()
    
    messages.success(request, 'Lesson deleted successfully!')
    
    # Redirect based on user role
    if user_profile.role == 'admin':
        return redirect('admin_courses')
    else:
        return redirect('manager_courses')

@login_required
def course_details(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    
    # Check if user is enrolled in this course
    course.is_enrolled = Enrollment.objects.filter(
        user=request.user,
        course=course
    ).exists()
    
    # Get completion status for lessons if user is enrolled
    if course.is_enrolled:
        # Get the course progress for this user
        progress = CourseProgress.objects.filter(
            user=request.user,
            course=course
        ).first()
        
        # Get all lessons for the course
        lessons = course.lessons.all().order_by('order')
        
        # If there's progress, mark completed lessons
        if progress:
            completed_lessons = progress.completed_lessons.all().values_list('id', flat=True)
            
            # Mark completed lessons
            for lesson in lessons:
                lesson.is_completed = lesson.id in completed_lessons
            
            # Calculate completion rate
            total_lessons = lessons.count()
            completed_count = len(completed_lessons)
            course.lesson_completion_rate = int((completed_count / total_lessons * 100) if total_lessons > 0 else 0)
            
            # Get or update StudentProgress for overall tracking
            student_progress, _ = StudentProgress.objects.get_or_create(
                user=request.user,
                course=course
            )
            student_progress.progress_percentage = course.lesson_completion_rate
            student_progress.save()
        else:
            # If no progress exists yet, initialize with all lessons as incomplete
            for lesson in lessons:
                lesson.is_completed = False
            course.lesson_completion_rate = 0
        
        course.lessons_list = lessons
    
    context = {
        'course': course,
        'user_profile': request.user.userprofile,
    }
    return render(request, 'employee/course_details.html', context)

@login_required
def continue_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    enrollment = get_object_or_404(Enrollment, user=request.user, course=course)
    
    # Get the course progress
    course_progress = CourseProgress.objects.filter(
        user=request.user,
        course=course
    ).first()
    
    if course_progress:
        # Get completed lesson IDs
        completed_lessons = course_progress.completed_lessons.all().values_list('id', flat=True)
        # Find the first incomplete lesson
        next_lesson = course.lessons.exclude(id__in=completed_lessons).order_by('order').first()
    else:
        # If no progress exists, start with the first lesson
        next_lesson = course.lessons.order_by('order').first()
    
    if next_lesson:
        return redirect('view_lesson', lesson_id=next_lesson.id)
    else:
        messages.warning(request, 'This course has no lessons yet.')
        return redirect('course_details', course_id=course_id)

@login_required
def view_lesson(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)
    course = lesson.course
    
    # Check if user is enrolled in the course
    enrollment = get_object_or_404(Enrollment, user=request.user, course=course)
    
    # Get or create progress record
    progress, created = CourseProgress.objects.get_or_create(
        user=request.user,
        course=course
    )
    
    # Get all lessons for the course in order
    all_lessons = list(course.lessons.order_by('order'))
    current_lesson_index = all_lessons.index(lesson)
    
    # Get completed lessons
    completed_lessons = progress.completed_lessons.all()
    
    # Check if previous lessons are completed
    can_mark_complete = True
    if current_lesson_index > 0:
        previous_lesson = all_lessons[current_lesson_index - 1]
        if previous_lesson not in completed_lessons:
            can_mark_complete = False
    
    # Handle lesson completion
    if request.method == 'POST' and 'mark_complete' in request.POST:
        if can_mark_complete and lesson not in completed_lessons:
            progress.completed_lessons.add(lesson)
            
            # Calculate new progress percentage
            total_lessons = course.lessons.count()
            completed_count = progress.completed_lessons.count()
            
            # Update StudentProgress for overall tracking
            student_progress, _ = StudentProgress.objects.get_or_create(
                user=request.user,
                course=course
            )
            student_progress.progress_percentage = (completed_count / total_lessons) * 100
            student_progress.save()
            
            # Update enrollment progress
            enrollment.progress = student_progress.progress_percentage
            if student_progress.progress_percentage == 100:
                enrollment.completed = True
                enrollment.completed_at = timezone.now()
            enrollment.save()
            
            messages.success(request, 'Lesson marked as complete!')
            
            # If there are more lessons, redirect to the next one
            if current_lesson_index < len(all_lessons) - 1:
                next_lesson = all_lessons[current_lesson_index + 1]
                return redirect('view_lesson', lesson_id=next_lesson.id)
            else:
                messages.success(request, 'Congratulations! You have completed the course!')
                return redirect('course_details', course_id=course.id)
    
    # Get content blocks in order
    content_blocks = lesson.content_blocks.order_by('order')
    
    # Add lessons_list to course object and mark completed lessons
    course.lessons_list = all_lessons
    for course_lesson in course.lessons_list:
        course_lesson.is_completed = course_lesson in completed_lessons
    
    context = {
        'lesson': lesson,
        'course': course,
        'content_blocks': content_blocks,
        'can_mark_complete': can_mark_complete,
        'is_completed': lesson in completed_lessons,
        'user_profile': request.user.userprofile,
        'progress': progress,
    }
    return render(request, 'employee/view_lesson.html', context)

@login_required
def export_users(request):
    # Check if the user is a manager
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.role != 'manager':
            messages.error(request, "Access denied. Manager privileges required.")
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect('home')

    # Get manager's courses
    manager_courses = Course.objects.filter(instructor=request.user)
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Users Progress"

    # Add headers
    headers = ['User', 'Email', 'Phone', 'Country', 'Status', 'Enrolled Courses', 'Completed Courses']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Add course columns
    course_col = len(headers) + 1
    for course in manager_courses:
        ws.cell(row=1, column=course_col, value=course.title)
        course_col += 1

    # Get all employees
    employees = User.objects.filter(userprofile__role='employee')
    
    # Add user data
    for row, employee in enumerate(employees, 2):
        # Basic user info
        ws.cell(row=row, column=1, value=employee.get_full_name())
        ws.cell(row=row, column=2, value=employee.email)
        ws.cell(row=row, column=3, value=employee.userprofile.phone)
        ws.cell(row=row, column=4, value=employee.userprofile.country)
        ws.cell(row=row, column=5, value='Active' if employee.is_active else 'Inactive')
        
        # Course counts
        enrolled_count = Enrollment.objects.filter(
            user=employee,
            course__in=manager_courses
        ).count()
        completed_count = Enrollment.objects.filter(
            user=employee,
            course__in=manager_courses,
            completed=True
        ).count()
        
        ws.cell(row=row, column=6, value=enrolled_count)
        ws.cell(row=row, column=7, value=completed_count)
        
        # Course progress
        course_col = len(headers) + 1
        for course in manager_courses:
            progress = CourseProgress.objects.filter(
                user=employee,
                course=course
            ).first()
            
            if progress:
                total_lessons = course.lessons.count()
                if total_lessons > 0:
                    completed_lessons = progress.completed_lessons.count()
                    progress_percentage = (completed_lessons / total_lessons) * 100
                    ws.cell(row=row, column=course_col, value=f"{progress_percentage:.1f}%")
                else:
                    ws.cell(row=row, column=course_col, value="0%")
            else:
                ws.cell(row=row, column=course_col, value="Not Enrolled")
            course_col += 1

    # Create the HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=users_progress.xlsx'
    
    # Save the workbook
    wb.save(response)
    
    return response

@login_required
def toggle_user_status(request, user_id):
    # Check if the user is a manager
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        if user_profile.role != 'manager':
            messages.error(request, "Access denied. Manager privileges required.")
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect('home')

    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            if user.userprofile.role == 'employee':
                user.is_active = not user.is_active
                user.save()
                messages.success(request, f"User {user.get_full_name()} has been {'activated' if user.is_active else 'deactivated'}.")
            else:
                messages.error(request, "You can only toggle status for employees.")
        except User.DoesNotExist:
            messages.error(request, "User not found.")
    
    return redirect('users')

@user_passes_test(is_admin, login_url='login')
def admin_dashboard(request):
    # Get total users count
    total_users = UserProfile.objects.count()
    
    # Get user growth (new users in last 30 days)
    thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
    user_growth = UserProfile.objects.filter(user__date_joined__gte=thirty_days_ago).count()
    
    # Get total courses count
    total_courses = Course.objects.count()
    
    # Get course growth (new courses in last 30 days)
    course_growth = Course.objects.filter(created_at__gte=thirty_days_ago).count()
    
    # Get active managers count
    active_managers = UserProfile.objects.filter(role='manager', user__is_active=True).count()
    
    # Get course status counts
    active_courses = Course.objects.filter(is_active=True).count()
    featured_courses = Course.objects.filter(is_featured=True).count()
    
    # Calculate course completion rate - average across all courses and users
    total_completion_percentage = 0
    total_enrollments = 0
    
    # Get all courses
    courses = Course.objects.all()
    for course in courses:
        # Get all enrollments for this course
        enrollments = Enrollment.objects.filter(course=course)
        for enrollment in enrollments:
            progress = CourseProgress.objects.filter(
                user=enrollment.user,
                course=course
            ).first()
            
            if progress and course.lessons.count() > 0:
                completed_lessons = progress.completed_lessons.count()
                total_lessons = course.lessons.count()
                completion_percentage = (completed_lessons / total_lessons) * 100
                total_completion_percentage += completion_percentage
                total_enrollments += 1
    
    # Calculate average completion rate
    course_completion_rate = (total_completion_percentage / total_enrollments) if total_enrollments > 0 else 0
    
    # Get recent activities
    recent_activities = Activity.objects.select_related('user', 'course').order_by('-created_at')[:10]
    
    # Get latest users
    latest_users = UserProfile.objects.select_related('user').order_by('-user__date_joined')[:5]
    
    context = {
        'total_users': total_users,
        'user_growth': user_growth,
        'total_courses': total_courses,
        'course_growth': course_growth,
        'active_managers': active_managers,
        'active_courses': active_courses,
        'featured_courses': featured_courses,
        'course_completion_rate': round(course_completion_rate, 1),
        'recent_activities': recent_activities,
        'latest_users': latest_users,
    }
    
    return render(request, 'admin/dashboard.html', context)

@user_passes_test(is_admin)
def admin_users(request):
    # Get query parameters
    search_query = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort', 'name')  # Default sort by name
    sort_order = request.GET.get('order', 'asc')  # Default ascending order
    page = request.GET.get('page', 1)

    # Base queryset
    users = UserProfile.objects.select_related('user').all()

    # Apply search filter
    if search_query:
        users = users.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

    # Apply role filter
    if role_filter:
        users = users.filter(role=role_filter)

    # Apply status filter
    if status_filter:
        is_active = status_filter == 'active'
        users = users.filter(user__is_active=is_active)

    # Apply sorting
    sort_field = {
        'name': 'user__first_name',
        'email': 'user__email',
        'role': 'role',
        'status': 'user__is_active',
        'joined': 'user__date_joined'
    }.get(sort_by, 'user__first_name')

    if sort_order == 'desc':
        sort_field = f'-{sort_field}'
    users = users.order_by(sort_field)

    # Pagination
    paginator = Paginator(users, 10)  # Show 10 users per page
    try:
        users_page = paginator.page(page)
    except PageNotAnInteger:
        users_page = paginator.page(1)
    except EmptyPage:
        users_page = paginator.page(paginator.num_pages)

    # Get quick stats
    total_users = UserProfile.objects.count()
    active_users = UserProfile.objects.filter(user__is_active=True).count()
    inactive_users = total_users - active_users
    
    # Calculate percentages
    active_percentage = round((active_users / total_users * 100) if total_users > 0 else 0)
    inactive_percentage = round((inactive_users / total_users * 100) if total_users > 0 else 0)

    # Calculate user growth
    thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
    new_users = UserProfile.objects.filter(user__date_joined__gte=thirty_days_ago).count()
    user_growth = round((new_users / total_users * 100) if total_users > 0 else 0)

    # Calculate course completion rate
    total_completion_percentage = 0
    total_enrollments = 0
    completed_courses = 0
    
    enrollments = Enrollment.objects.all()
    for enrollment in enrollments:
        if enrollment.completed:
            completed_courses += 1
        progress = CourseProgress.objects.filter(
            user=enrollment.user,
            course=enrollment.course
        ).first()
        
        if progress and enrollment.course.lessons.count() > 0:
            completed_lessons = progress.completed_lessons.count()
            total_lessons = enrollment.course.lessons.count()
            completion_percentage = (completed_lessons / total_lessons) * 100
            total_completion_percentage += completion_percentage
            total_enrollments += 1
    
    course_completion_rate = round(total_completion_percentage / total_enrollments if total_enrollments > 0 else 0)

    context = {
        'users': users_page,
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'active_percentage': active_percentage,
        'inactive_percentage': inactive_percentage,
        'user_growth': user_growth,
        'course_completion_rate': course_completion_rate,
        'completed_courses': completed_courses,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'role_choices': UserProfile.ROLE_CHOICES,
    }
    return render(request, 'admin/users.html', context)

@user_passes_test(is_admin)
def update_user_role(request, user_id):
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            user_profile = user.userprofile
            new_role = request.POST.get('role')
            if new_role in dict(UserProfile.ROLE_CHOICES):
                user_profile.role = new_role
                user_profile.save()
                messages.success(request, f'Role updated successfully for {user.get_full_name()}')
            else:
                messages.error(request, 'Invalid role selected')
        except User.DoesNotExist:
            messages.error(request, 'User not found')
    return redirect('admin_users')

@login_required
def delete_user(request, user_id):
    if not request.user.userprofile.role == 'admin':
        messages.error(request, 'You do not have permission to delete users.')
        return redirect('admin_dashboard')
    
    try:
        user = User.objects.get(id=user_id)
        # Delete the user (this will cascade delete the UserProfile)
        user.delete()
        messages.success(request, f'User {user.get_full_name()} has been deleted successfully.')
    except User.DoesNotExist:
        messages.error(request, 'User not found.')
    except Exception as e:
        messages.error(request, f'Error deleting user: {str(e)}')
    
    return redirect('admin_users')

@user_passes_test(is_admin)
def add_user(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Create User instance
                user = User.objects.create_user(
                    username=request.POST['email'],  # Using email as username
                    email=request.POST['email'],
                    password=request.POST.get('password', User.objects.make_random_password()),  # Generate random password if not provided
                    first_name=request.POST['first_name'],
                    last_name=request.POST['last_name'],
                )
                
                # Update or create UserProfile
                user_profile, created = UserProfile.objects.update_or_create(
                    user=user,
                    defaults={
                        'role': request.POST['role'],
                        'phone': request.POST.get('phone', ''),
                        'gender': request.POST.get('gender', 'other'),
                    }
                )

                if 'photo' in request.FILES:
                    user_profile.profile_picture = request.FILES['photo']
                    user_profile.save()

                messages.success(request, 'User created successfully!')
                return redirect('admin_users')
        except Exception as e:
            messages.error(request, f'Error creating user: {str(e)}')
            return redirect('add_user')

    # Get statistics for the quick stats card
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    
    # Get recent users for the table
    recent_users = User.objects.select_related('userprofile').order_by('-date_joined')[:5]

    context = {
        'total_users': total_users,
        'active_users': active_users,
        'recent_users': recent_users,
    }
    return render(request, 'admin/add_user.html', context)

@user_passes_test(is_admin)
def admin_courses(request):
    if request.user.userprofile.role != 'admin':
        return redirect('home')
    
    # Get all courses
    courses = Course.objects.all()
    
    # Apply filters
    search_query = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    status = request.GET.get('status', '')
    sort_by = request.GET.get('sort', '')
    sort_order = request.GET.get('order', 'asc')
    
    if search_query:
        courses = courses.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(instructor__first_name__icontains=search_query) |
            Q(instructor__last_name__icontains=search_query)
        )
    
    if category_id:
        courses = courses.filter(category_id=category_id)
    
    if status == 'active':
        courses = courses.filter(is_active=True)
    elif status == 'inactive':
        courses = courses.filter(is_active=False)
    
    # Apply sorting
    if sort_by:
        if sort_by == 'title':
            courses = courses.order_by(f"{'-' if sort_order == 'desc' else ''}title")
        elif sort_by == 'instructor':
            courses = courses.order_by(f"{'-' if sort_order == 'desc' else ''}instructor__first_name")
        elif sort_by == 'category':
            courses = courses.order_by(f"{'-' if sort_order == 'desc' else ''}category__name")
        elif sort_by == 'enrollments':
            courses = courses.annotate(
                enrolled_count=Count('enrollments')
            ).order_by(f"{'-' if sort_order == 'desc' else ''}enrolled_count")
        elif sort_by == 'completion':
            courses = courses.annotate(
                completion_rate=Avg('enrollments__progress')
            ).order_by(f"{'-' if sort_order == 'desc' else ''}completion_rate")
        elif sort_by == 'status':
            courses = courses.order_by(f"{'-' if sort_order == 'desc' else ''}is_active")
    else:
        # Default sorting by creation date
        courses = courses.order_by('-created_at')
    
    # Calculate statistics
    total_courses = Course.objects.count()
    active_courses = Course.objects.filter(is_active=True).count()
    active_percentage = (active_courses / total_courses * 100) if total_courses > 0 else 0
    
    # Calculate course growth
    last_month = timezone.now() - timezone.timedelta(days=30)
    new_courses = Course.objects.filter(created_at__gte=last_month).count()
    course_growth = (new_courses / total_courses * 100) if total_courses > 0 else 0
    
    # Calculate enrollment statistics
    total_enrollments = Enrollment.objects.count()
    completed_enrollments = Enrollment.objects.filter(progress=100).count()
    completion_rate = (completed_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0
    
    # Calculate enrollment growth
    new_enrollments = Enrollment.objects.filter(enrolled_at__gte=last_month).count()
    enrollment_growth = (new_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0
    
    # Pagination
    paginator = Paginator(courses, 10)
    page = request.GET.get('page')
    courses = paginator.get_page(page)
    
    context = {
        'courses': courses,
        'categories': CourseCategory.objects.all(),
        'total_courses': total_courses,
        'active_courses': active_courses,
        'active_percentage': round(active_percentage, 1),
        'course_growth': round(course_growth, 1),
        'total_enrollments': total_enrollments,
        'enrollment_growth': round(enrollment_growth, 1),
        'completion_rate': round(completion_rate, 1),
        'completed_courses': completed_enrollments,
        'sort_by': sort_by,
        'sort_order': sort_order,
    }
    
    return render(request, 'admin/courses.html', context)

@user_passes_test(is_admin)
def admin_categories(request):
    categories = CourseCategory.objects.all()
    return render(request, 'admin/categories.html', {'categories': categories})

@user_passes_test(is_admin)
def admin_reports(request):
    # Add report generation logic here
    return render(request, 'admin/reports.html')

@user_passes_test(is_admin)
def admin_settings(request):
    # Add settings management logic here
    return render(request, 'admin/settings.html')

def chat_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    messages = ChatMessage.objects.filter(user=request.user).order_by('timestamp')
    return render(request, 'chat.html', {'messages': messages})

def send_message(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'User not authenticated'}, status=401)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method is allowed'}, status=405)
    
    try:
        message = request.POST.get('message')
        print(f"Received message: {message}")  # Debug log
        
        if not message:
            return JsonResponse({'error': 'Message field is missing'}, status=400)
        
        message = message.strip()
        if not message:
            return JsonResponse({'error': 'Message cannot be empty'}, status=400)
        
        # Save user message
        user_message = ChatMessage.objects.create(
            user=request.user,
            message=message,
            is_bot=False
        )
        
        # Get chat history
        chat_history = ChatMessage.objects.filter(user=request.user).order_by('-timestamp')[:5]
        
        # Get bot response
        try:
            bot_response = get_bot_response(message, chat_history)
            if not bot_response:
                raise ValueError("Empty response from bot")
                
        except Exception as e:
            print(f"Error getting bot response: {str(e)}")  # Log the error
            return JsonResponse({
                'error': 'Failed to get response from chat service',
                'details': str(e),
                'user_message': {
                    'message': user_message.message,
                    'timestamp': user_message.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                }
            }, status=500)
        
        # Save bot response
        bot_message = ChatMessage.objects.create(
            user=request.user,
            message=bot_response,
            is_bot=True
        )
        
        return JsonResponse({
            'user_message': {
                'message': user_message.message,
                'timestamp': user_message.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            },
            'bot_message': {
                'message': bot_message.message,
                'timestamp': bot_message.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            }
        })
        
    except Exception as e:
        print(f"Error in send_message view: {str(e)}")  # Log the error
        return JsonResponse({
            'error': 'An unexpected error occurred',
            'details': str(e)
        }, status=500)

@login_required
def manager_available_courses(request):
    user_profile = get_object_or_404(UserProfile, user=request.user)
    if user_profile.role != 'manager':
        messages.error(request, 'Access denied. Manager privileges required.')
        return redirect('home')
    
    # Get all active courses that the user is not enrolled in and didn't create
    enrolled_course_ids = Course.objects.filter(
        enrollments__user=request.user
    ).values_list('id', flat=True)
    
    created_course_ids = Course.objects.filter(
        instructor=request.user
    ).values_list('id', flat=True)
    
    # Get admin users
    admin_users = User.objects.filter(userprofile__role='admin')
    
    available_courses = Course.objects.filter(
        is_active=True,
        instructor__in=admin_users  # Only show courses created by admins
    ).exclude(
        id__in=enrolled_course_ids
    ).exclude(
        id__in=created_course_ids
    ).order_by('-created_at')
    
    return render(request, 'manager/available_courses.html', {
        'courses': available_courses,
        'user_profile': user_profile
    })

def contact_us(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        company = request.POST.get('company')
        message = request.POST.get('message')
        subject = f"Contact Us Message from {name}"
        full_message = f"""
        Name: {name}
        Email: {email}
        Phone: {phone}
        Company: {company}
        Message: {message}
        """
        try:
            send_mail(
                subject,
                full_message,
                settings.DEFAULT_FROM_EMAIL,
                ['growth.mate01@gmail.com'],
                fail_silently=False,
            )
            messages.success(request, "Your message has been sent successfully!")
        except Exception as e:
            messages.error(request, f"Failed to send message. Please try again later. Error: {str(e)}")
        return redirect('contact_us')
    return render(request, 'contact_us.html')

def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, 'No user found with that email address.')
            return redirect('forgot_password')
        # Generate OTP
        otp = ''.join(random.choices(string.digits, k=6))
        request.session['reset_email'] = email
        request.session['reset_otp'] = otp
        request.session.set_expiry(300)  # 5 minutes expiry
        # Send OTP email
        send_mail(
            'Password Reset OTP',
            f'Your OTP for password reset is: {otp}',
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )
        messages.success(request, 'An OTP has been sent to your email address.')
        return redirect('verify_reset_otp')
    return render(request, 'forgot_password.html')

def verify_reset_otp(request):
    if request.method == 'POST':
        entered_otp = request.POST.get('otp')
        stored_otp = request.session.get('reset_otp')
        if not stored_otp:
            messages.error(request, 'Session expired. Please try again.')
            return redirect('forgot_password')
        if entered_otp == stored_otp:
            request.session['otp_verified'] = True
            return redirect('reset_password')
        else:
            messages.error(request, 'Invalid OTP. Please try again.')
            return redirect('verify_reset_otp')
    return render(request, 'verify_reset_otp.html')

def reset_password(request):
    if not request.session.get('otp_verified'):
        messages.error(request, 'OTP verification required.')
        return redirect('forgot_password')
    if request.method == 'POST':
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        if password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect('reset_password')
        email = request.session.get('reset_email')
        try:
            user = User.objects.get(email=email)
            user.set_password(password)
            user.save()
            # Clean up session
            for key in ['reset_email', 'reset_otp', 'otp_verified']:
                request.session.pop(key, None)
            messages.success(request, 'Your password has been reset. You can now log in.')
            return redirect('login')
        except User.DoesNotExist:
            messages.error(request, 'User not found.')
            return redirect('forgot_password')
    return render(request, 'reset_password.html')